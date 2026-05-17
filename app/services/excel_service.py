import os

from openpyxl import Workbook
from openpyxl import load_workbook


FILE_NAME = "Invoices_Master.xlsx"


def save_excel(base_path, data):

    os.makedirs(
        base_path,
        exist_ok=True
    )

    file_path = os.path.join(
        base_path,
        FILE_NAME
    )

    company = data.get(
        "company",
        "Unknown"
    )

    sheet_name = company[:31]

    row = [

        data.get("invoice_no"),

        data.get("date"),

        data.get("company"),

        data.get("total"),

        data.get("vat"),

        data.get("currency"),

        data.get("source")

    ]

    # CREATE FILE

    if not os.path.exists(file_path):

        wb = Workbook()

        ws = wb.active

        ws.title = sheet_name

        ws.append([

            "Invoice No",

            "Date",

            "Company",

            "Total",

            "VAT",

            "Currency",

            "Source"

        ])

        ws.append(row)

        wb.save(file_path)

        print("EXCEL FILE CREATED")

        return True

    # LOAD EXISTING

    wb = load_workbook(file_path)

    # CREATE SHEET IF NOT EXISTS

    if sheet_name not in wb.sheetnames:

        ws = wb.create_sheet(
            title=sheet_name
        )

        ws.append([

            "Invoice No",

            "Date",

            "Company",

            "Total",

            "VAT",

            "Currency",

            "Source"

        ])

    else:

        ws = wb[sheet_name]

    # DUPLICATE CHECK

    for existing_row in ws.iter_rows(
        min_row=2,
        values_only=True
    ):

        if existing_row[0] == data.get("invoice_no"):

            print("DUPLICATE INVOICE SKIPPED")

            return False

    ws.append(row)

    wb.save(file_path)

    print("EXCEL UPDATED")

    return True